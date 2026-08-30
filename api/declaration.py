"""Module for generating monthly purchase/sale declaration Excel files.

The module parses a free‑text period description, retrieves purchase and sale
lines from the database using the `adaptation.db_adapter` abstraction,
formats dates, computes totals (HT, TVA, TTC) and writes a styled Excel
workbook. The result is saved in the ``declarations_generes`` directory and
a JSON summary is returned.
"""

import json
from decimal import Decimal

_json_dumps_orig = json.dumps
def _json_dumps_safe(obj, **kwargs):
    """Serialize ``obj`` to JSON, converting ``Decimal`` to ``float`` and other
    unsupported types to ``str``.
    """
    kwargs.setdefault("default", lambda o: float(o) if isinstance(o, Decimal) else str(o))
    return _json_dumps_orig(obj, **kwargs)
json.dumps = _json_dumps_safe

import re
import sys
from datetime import datetime
from pathlib import Path
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent))
from adaptation.db_adapter import table, col, get_connection

OUTPUT_DIR = Path("./declarations_generes")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

TVA_TAUX = 0.19

_MOIS_FR = {
    "janvier": 1, "jan": 1, "février": 2, "fevrier": 2, "fev": 2,
    "mars": 3, "avril": 4, "avr": 4, "mai": 5,
    "juin": 6, "juillet": 7, "juil": 7, "août": 8, "aout": 8,
    "septembre": 9, "sept": 9, "octobre": 10, "oct": 10,
    "novembre": 11, "nov": 11, "décembre": 12, "decembre": 12, "dec": 12,
}
_MOIS_NOMS = [
    "", "JANVIER", "FÉVRIER", "MARS", "AVRIL", "MAI", "JUIN",
    "JUILLET", "AOÛT", "SEPTEMBRE", "OCTOBRE", "NOVEMBRE", "DÉCEMBRE",
]

def _is_mssql() -> bool:
    """Return ``True`` if the environment variable ``DB_DRIVER`` is set to ``mssql``."""
    return os.getenv('DB_DRIVER', 'sqlite').lower() == 'mssql'

def _parse_mois_annee(texte: str) -> tuple[int, int]:
    """Extract month and year from free‑text ``texte``.
    
    If the year is not found, the current year is used.
    """
    texte = (texte or "").lower()
    mois = None
    for nom, num in _MOIS_FR.items():
        if re.search(rf"\b{nom}\b", texte):
            mois = num
            break
    if mois is None:
        m = re.search(r"\b(0?[1-9]|1[0-2])[/\-](\d{4})\b", texte)
        if m:
            return int(m.group(1)), int(m.group(2))
        raise ValueError("Mois introuvable dans la demande.")
    m_annee = re.search(r"\b(20\d{2})\b", texte)
    annee = int(m_annee.group(1)) if m_annee else datetime.now().year
    return mois, annee

def _connect():
    """Obtain a database connection using the ``adaptation.db_adapter`` helper."""
    return get_connection()

def _fmt_date(v) -> str:
    """Format a date value returned by the DB driver as ``JJ/MM/AAAA``.
    
    Handles both ``datetime`` objects (MSSQL) and ISO date strings (SQLite).
    """
    if hasattr(v, "strftime"):
        return v.strftime("%d/%m/%Y")
    return datetime.strptime(str(v), "%Y-%m-%d").strftime("%d/%m/%Y")

def _lignes_periode(conn, domaine: int, mois: int, annee: int, label_tiers: str):
    """Retrieve purchase (domaine=1) or sale (domaine=0) lines for the given period.
    
    Returns a list of dictionaries containing formatted date, reference,
    tier name, and monetary amounts (HT, TVA, TTC).
    """
    doc_entete_table = table("doc_entete")
    clients_table    = table("clients_fournisseurs")
    doc_ligne_table  = table("doc_ligne")

    piece_col   = col("doc_entete", "piece")
    date_col    = col("doc_entete", "date")
    type_col    = col("doc_entete", "type")
    domaine_col = col("doc_entete", "domaine")
    tiers_col   = col("doc_entete", "code_tiers")

    code_col = col("clients_fournisseurs", "code")
    nom_col  = col("clients_fournisseurs", "nom")

    piece_col_ligne = col("doc_ligne", "piece")
    qte_col         = col("doc_ligne", "qte")
    prix_col        = col("doc_ligne", "prix_unitaire")

    if _is_mssql():
        filtre_periode = f"YEAR(e.{date_col}) = ? AND MONTH(e.{date_col}) = ?"
        params = (annee, mois)
    else:
        mois_str = f"{annee:04d}-{mois:02d}"
        filtre_periode = f"STRFTIME('%Y-%m', e.{date_col}) = ?"
        params = (mois_str,)

    rows = conn.execute(f"""
        SELECT e.{piece_col} AS DO_Piece, e.{date_col} AS DO_Date, e.{tiers_col} AS CT_Num,
               COALESCE(c.{nom_col}, e.{tiers_col}) AS tiers,
               COALESCE(SUM(l.{qte_col} * l.{prix_col}), 0) AS ht
        FROM {doc_entete_table} e
        LEFT JOIN {clients_table}  c ON e.{tiers_col} = c.{code_col}
        LEFT JOIN {doc_ligne_table} l ON e.{piece_col} = l.{piece_col_ligne}
        WHERE e.{type_col} = {6 if domaine == 0 else 16} AND e.{domaine_col} = {domaine}
          AND {filtre_periode}
        GROUP BY e.{piece_col}, e.{date_col}, e.{tiers_col}, c.{nom_col}
        ORDER BY e.{date_col}
    """, params).fetchall()

    lignes = []
    for r in rows:
        ht = round(float(r["ht"] or 0.0), 2)
        tva = round(ht * TVA_TAUX, 2)
        ttc = round(ht + tva, 2)
        lignes.append({
            "date": _fmt_date(r["DO_Date"]),
            "ref": r["DO_Piece"],
            "tiers": r["tiers"],
            "ht": ht, "tva": tva, "ttc": ttc,
        })
    return lignes

FILL_TITRE = PatternFill("solid", fgColor="4BACC6")
FILL_ENTETE_SECTION = PatternFill("solid", fgColor="D9D9D9")
FILL_HEADER_COL = PatternFill("solid", fgColor="D9D9D9")
FILL_TOTAL = PatternFill("solid", fgColor="FBD4B4")
FONT_TITRE = Font(bold=True, color="FFFFFF", size=13)
FONT_ENTETE = Font(bold=True)
FONT_TOTAL = Font(bold=True)
BORDURE = Border(*(Side(style="thin", color="808080"),) * 4)
CENTRE = Alignment(horizontal="center", vertical="center")

def _bloc(ws, col_debut: int, titre: str, lignes: list[dict], label_tiers: str):
    """Write a purchase or sale block into ``ws`` starting at column ``col_debut``.
    
    Returns a tuple ``(total_ht, total_tva, total_ttc)`` for the block.
    """
    c0, c1, c2, c3, c4, c5 = range(col_debut, col_debut + 6)

    ws.merge_cells(start_row=3, start_column=c0, end_row=3, end_column=c5)
    cell = ws.cell(row=3, column=c0, value=titre)
    cell.font = FONT_ENTETE
    cell.alignment = CENTRE
    cell.fill = FILL_ENTETE_SECTION

    entetes = ["Date", "Référence", label_tiers, "H.T", "TVA", "TTC"]
    for i, txt in enumerate(entetes):
        cell = ws.cell(row=4, column=col_debut + i, value=txt)
        cell.font = FONT_ENTETE
        cell.fill = FILL_HEADER_COL
        cell.alignment = CENTRE
        cell.border = BORDURE

    r = 5
    total_ht = total_tva = total_ttc = 0.0
    for ligne in lignes:
        ws.cell(row=r, column=c0, value=ligne["date"]).border = BORDURE
        ws.cell(row=r, column=c1, value=ligne["ref"]).border = BORDURE
        ws.cell(row=r, column=c2, value=ligne["tiers"]).border = BORDURE
        ws.cell(row=r, column=c3, value=ligne["ht"]).border = BORDURE
        ws.cell(row=r, column=c4, value=ligne["tva"]).border = BORDURE
        ws.cell(row=r, column=c5, value=ligne["ttc"]).border = BORDURE
        for cc in (c3, c4, c5):
            ws.cell(row=r, column=cc).number_format = "0.00"
        total_ht  += ligne["ht"]
        total_tva += ligne["tva"]
        total_ttc += ligne["ttc"]
        r += 1

    ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c0, value="TOTAL")
    cell.font = FONT_TOTAL
    cell.fill = FILL_TOTAL
    cell.alignment = CENTRE
    for cc, val in ((c3, total_ht), (c4, total_tva), (c5, total_ttc)):
        cell = ws.cell(row=r, column=cc, value=round(val, 2))
        cell.font = FONT_TOTAL
        cell.fill = FILL_TOTAL
        cell.number_format = "0.00"
    for cc in range(c0, c5 + 1):
        ws.cell(row=r, column=cc).border = BORDURE

    for i in range(6):
        ws.column_dimensions[get_column_letter(col_debut + i)].width = 14

    return total_ht, total_tva, total_ttc

def generer_declaration_mensuelle_excel(texte_periode: str) -> str:
    """Generate the monthly declaration Excel file for the period described by ``texte_periode``.
    
    Returns a JSON string summarising the operation, including status, month,
    year, file path and aggregated totals for purchases and sales.
    """
    mois, annee = _parse_mois_annee(texte_periode)
    conn = _connect()
    try:
        achats = _lignes_periode(conn, 1, mois, annee, "Fournisseur")
        ventes = _lignes_periode(conn, 0, mois, annee, "Client")
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Déclaration"
    ws.sheet_view.showGridLines = False

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    titre_cell =